from asyncio.log import logger
from ctypes import sizeof
import sys
from difflib import get_close_matches
from pathlib import Path

import openpyxl
from flask import abort, send_file
from flask import current_app as app
from opencve.commands import create_category, ensure_config, error, info
from opencve.controllers.base import BaseController
from opencve.extensions import db
from opencve.models.categories import Category
from opencve.models.products import Product
from opencve.models.vendors import Vendor
from opencve.models.cve import Cve
from opencve.models.alerts import Alert
from opencve.utils import get_categories_letters
from sqlalchemy.exc import IntegrityError
import datetime
from openpyxl import Workbook
from flask_user import current_user
from sqlalchemy import and_, or_





class CategoryController(BaseController):
    model = Category
    order = Category.name.asc()
    per_page_param = "CATEGORYS_PER_PAGE"
    schema = {
        "letter": {"type": str},
        "search": {"type": str},
        "category_name": {"type": str},
    }

    @classmethod
    def build_query(cls, args):
        letter = args.get("letter")

        query = cls.model.query

        # Search by term
        if args.get("search"):
            search = args.get("search").lower().replace(
                "%", "").replace("_", "")
            query = query.filter(cls.model.name.like("%{}%".format(search)))

        # Search by letter
        if letter:
            if letter not in get_categories_letters():
                abort(404)

            query = query.filter(cls.model.name.like("{}%".format(letter)))

        return query, {}


VENDORS = []
PRODUCTS = []


def dehumanize_filter(name):
    """Reverses opencve.context._humanize_filter"""
    return "_".join(map(lambda x: x.lower(), name.split(" ")))


def find_db_name(Model, name):
    """Uses difflib.get_close_matches to find the closest match to name in the Model specified
    Made to work only with Vendor and Product"""
    try:
        name = dehumanize_filter(name)
        info("[FIND_DB_NAME] Searching for name : " + name)

        human_names = VENDORS if Model is Vendor else PRODUCTS

        db_name = get_close_matches(name, human_names, n=1, cutoff=0.6)
        if len(db_name) > 0:
            db_name = db_name[0]
        else:
            raise LookupError

        info("[FIND_DB_NAME] Found db_name : " + db_name)
        return db_name
    except LookupError:
        error(name + " not found in db names")
        return ""
    except Exception as e:
        error(e)
        return ""


def find_product_id(vendor, product, version):
    """Returns a product id or None if not found
    Uses find_db_name, a closest match searching algorithm"""
    vendor_query = Vendor.query.filter_by(
        name=find_db_name(Vendor, vendor)).first()
    if vendor_query is None:
        return info(f"[FIND_PRODUCT_ID] {vendor} does not exists in Vendors")
    vendor_query = vendor_query.id

    product_query = Product.query.filter_by(name=find_db_name(
        Product, version), vendor_id=vendor_query).first()
    if product_query is None:
        return info(f"[FIND_PRODUCT_ID] {product} does not exists in Products")

    return product_query.id


def add_product(category, vendor, product, version, tag):
    """Add Product to category"""
    category_query = Category.query.filter_by(name=category.name).first().id
    if not category_query:
        return info(f"[ADD_PRODUCT] {category} does not exists in Categories")

    if tag is None:
        product_id = find_product_id(vendor, product, version)
        if product_id is None:
            return
    else:
        product_query = Product.query.filter_by(name=str(tag)).first()
        if product_query is None:
            return info(f"[ADD_PRODUCT] {tag} does not exists in Products")
        product_id = product_query.id

    # FOR TESTING ONLY, NEED TO CHECK THE INPUT OR CHANGE THE USED METHOD //TODO
    existance = db.session.execute("SELECT 1 FROM categories_products WHERE category_id=('"+str(
        category_query)+"') and product_id=('"+str(product_id)+"')").first()
    if existance:
        return info(f"[ADD_PRODUCT] {product} already exists in category {category}")
    db.session.execute("INSERT INTO categories_products(category_id,product_id) VALUES (('" +
                       str(category_query)+"'),('"+str(product_id)+"'))")
    db.session.commit()


def create_category(name):
    """Create a Category."""
    name = str(name).lower()
    if Category.query.filter_by(name=name).first():
        return -1

    category = Category(
        name=name,
    )
    db.session.add(category)
    try:
        db.session.commit()
    except IntegrityError as e:
        error(e)
        return -1


def edit_category_name(category, name):
    """Edit the specified Category name."""
    name = str(name).lower()
    if Category.query.filter_by(name=name).first():
        return -1
    category.name = name
    try:
        db.session.commit()
    except IntegrityError as e:
        error(e)
        return -1


def delete_category(category):
    """Delete the specified Category."""
    try:
        db.session.delete(category)
        db.session.commit()
    except IntegrityError as e:
        error(e)
        return -1


def read_excel(category, xlsx_file):
    """Read an xlsx file and expects from it to have columns named vendor, product, version and tag
    Those names shall be found in the first three rows"""
    global PRODUCTS
    global VENDORS
    VENDORS = [x.name for x in Vendor.query.all()]
    PRODUCTS = [x.name for x in Product.query.all()]
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active
    data = []
    vendor_col = None
    product_col = None
    version_col = None
    tag_col = None
    min_value_index = None
    max_value_index = sheet.max_row

    for row in range(1, 3):  # To be sure, we check the 3 firsts rows
        tag_col = None
        for col in range(1, sheet.max_column):
            if str(sheet[row][col].value).lower() == "vendor":
                vendor_col = col
                min_value_index = row+1
            elif str(sheet[row][col].value).lower() == "product":
                product_col = col
            elif str(sheet[row][col].value).lower() == "version":
                version_col = col
            elif str(sheet[row][col].value).lower() == "tag":
                tag_col = col

    if vendor_col is None or product_col is None or version_col is None:
        error("[READ_EXCEL] Column format is not good")
        return -1

    # For every row, we check if the value is already in the data list
    for i in range(min_value_index, max_value_index):
        if sheet[i][vendor_col].value != None and sheet[i][product_col].value != None and sheet[i][version_col].value != None:
            d = str(sheet[i][vendor_col].value).lower()+":"+str(sheet[i][product_col].value).lower()+":"+str(sheet[i][version_col].value).lower()
            if tag_col != None:
                tag = sheet[i][tag_col].value
            else:
                tag = None
            if (d, tag) not in data:
                data.append((d, tag))
    ite = 0
    for (i, tag) in data:
        ite += 1
        # For debug purpose
        # info("[READ_EXCEL] Iteration : " + str(ite))
        # info("[READ_EXCEL] vendor :" + str(i.split(":")[0]))
        # info("[READ_EXCEL] product :" + str(i.split(":")[1]))
        # info("[READ_EXCEL] version :" + str(i.split(":")[2]))
        add_product(category, str(i.split(":")[0]), str(
            i.split(":")[1]), str(i.split(":")[2]), tag)

    # Empty the arrays
    VENDORS = []
    PRODUCTS = []


def generateCategoryReport(category, period):
    """Generate a report for the specified category"""
    # UPLOAD_FOLDER = '/app/venv/lib/python3.7/site-packages/opencve/data/'
    # Create a xlsx file with the category name
    # The file is named after the category and the datetime
    file_name = str(category.name) + "_" + str(datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S"))
    wb = Workbook()
    ws = wb.active

    # Print the products of the category and the number of CVEs associated to them
    date = datetime.datetime.now() - datetime.timedelta(days=period)
    ws["A1"] = "Product"
    ws["B1"] = "Vendor"
    ws["C1"] = "Number of CVEs since" + str(date.strftime("%Y-%m-%d"))
    ws["D1"] = "CVSS2 Score mean"
    ws["E1"] = "CVSS3 Score mean"
    ws["F1"] = "Number of critical CVEs (above 7.5)"
    i = 2
    total = 0
    cves = []

    for product in category.products:
        ws['A'+str(i)] = product.name
        ws['B'+str(i)] = product.vendor.name
        count = 0
        cveQuery = Cve.query.filter(
            and_(
                or_(
                    # Cve.vendors.contains([product.vendor.name]) if product.vendor else None, # For the moment, the count is also based on the vendor
                    Cve.vendors.contains([product.vendor.name+'$PRODUCT$'+product.name]) if product else None,
                ),
                Cve.updated_at >= date,
            )
        )

        cves += cveQuery
        count = cveQuery.count()
        total += count
        ws['C'+str(i)] = count
        i += 1
    ws['C'+str(i)] = total


    ws = wb.create_sheet("Vendors")
    ws["A1"] = "Vendor"
    ws["B1"] = "Number of CVEs since" + str(date.strftime("%Y-%m-%d"))
    ws["C1"] = "CVSS2 Score mean"
    ws["D1"] = "CVSS3 Score mean"
    ws["E1"] = "Number of critical CVEs (above 7.5)"
    i = 2
    total = 0

    for vendor in category.vendors:
        ws['A'+str(i)] = vendor.name
        count = 0
        cveQuery = Cve.query.filter(
            and_(
                or_(
                    Cve.vendors.contains([vendor.name]) if product.vendor else None, # For the moment, the count is also based on the vendor
                    # Cve.vendors.contains([product.vendor.name+'$PRODUCT$'+product.name]) if product else None,
                ),
                Cve.updated_at >= date,
            )
        )
        
        cves += cveQuery
        count = cveQuery.count()
        total += count
        ws['C'+str(i)] = count
        i += 1
    ws['C'+str(i)] = total


    # Print all the CVEs associated to the category
    # TODO: FIX. Some CVEs are selected, even if they are not associated to the category
    ws = wb.create_sheet("CVEs")
    ws['A1'] = "Creation date"
    ws['B1'] = "Last update"
    ws['C1'] = "CVE ID"
    ws['D1'] = "Vendor"
    ws['E1'] = "Product"
    ws['F1'] = "Description"
    ws['G1'] = "References"
    ws['H1'] = "CWE"
    ws['I1'] = "CVSS2"
    ws['J1'] = "CVSS3"
    total = 0
    i = 2

    for cve in cves:
        ws['A'+str(i)] = cve.created_at.strftime("%Y-%m-%d")
        ws['B'+str(i)] = cve.updated_at.strftime("%Y-%m-%d")
        ws['C'+str(i)] = cve.cve_id
        CVEvendors = []
        CVEproducts = []
        tab = str(cve.vendors).split('\'')

        for j in range(len(cve.vendors)):
            if tab[j] != '[' and tab[j] != ']' and ',' not in tab[j]:
                if '$' not in tab[j]: # TODO : Maybe change it to only display followed impacted vendors, or make it more easy to understand 
                    CVEvendors.append(tab[j])
                else: # TODO : Maybe change it to only display followed impacted products, or make it more easy to understand 
                    CVEproducts.append(str(cve.vendors).split('\'')[j].split('$')[2])
        ws['D'+str(i)] = str(CVEvendors)
        ws['E'+str(i)] = str(CVEproducts)
        # ws['F'+str(i)] = cve.summary
        # ws['G'+str(i)] = cve.references
        # ws['H'+str(i)] = cve.cwe
        # ws['I'+str(i)] = cve.cvss2
        # ws['J'+str(i)] = cve.cvss3
        i += 1

    wb.save("/"+file_name+".xlsx")
    return send_file("/"+file_name+".xlsx", as_attachment=True, attachment_filename=file_name+".xlsx")